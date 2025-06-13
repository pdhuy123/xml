import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

Dep = ["Dependent"]
Indep = ["Independent"]

def convert(value, unit):
    if pd.isna(value):
        return value
    try:
        value = float(value)
        unit = str(unit).strip()
        if unit == "GPa":
            return value * 1e9
        elif unit == "MPa":
            return value * 1e6
        elif unit == "kg/m^3":
            return value / 1000  # => g/cm^3
        elif unit == "kg/cm^3":
            return value * 1000  # => g/cm^3
        elif unit in ["1/K", "1/C"]:
            return value
        return value
    except:
        return value
    
def convert_unit(values, unit):
    lst = []
    for value in str(values).split(","):
        lst.append(str(convert(value, unit)))
    return ",".join(lst), len(lst)

def excel_to_ansys_fixed(input_file, output_file):
    df = pd.read_excel(input_file)

    # RGB color extraction
    df = pd.read_excel(r"D:\clf\xml\xml.xlsx")
    wb = load_workbook(r"D:\clf\xml\xml.xlsx", data_only=True)
    ws = wb.active
    colors = []
    for row in range(2, ws.max_row + 1): 
        cell = ws[f"B{row}"]
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb 
            colors.append((int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)))
        else:
            colors.append(None)
    df['RGB'] = colors
    df[['Red', 'Green', 'Blue']] = pd.DataFrame(df['RGB'].tolist(), index=df.index)
    # XML generation
    pa_set = set()
    pr_set = set()
    eng = ET.Element('EngineeringData', version="23.2.0.230",
                        versiondate=datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"))
    ET.SubElement(eng, 'Notes')
    materials = ET.SubElement(eng, 'Materials')
    matml = ET.SubElement(materials, 'MatML_Doc')

    for idx, row in df.iterrows():
        mat = ET.SubElement(matml, 'Material')
        bulk = ET.SubElement(mat, 'BulkDetails')
        ET.SubElement(bulk, 'Name').text = str(df.get('Name')[idx])

        # pr0: Color
        if any(pd.notna(row.get(c)) for c in ['Red', 'Green', 'Blue']):
            pr0 = ET.SubElement(bulk, 'PropertyData', property="pr0")
            ET.SubElement(pr0, 'Data', format='string').text = '-'
            for pid, col in zip(["pa0", "pa1", "pa2"], ['Red', 'Green', 'Blue']):
                val = row.get(col)
                fmt = 'string' if isinstance(val, str) else 'float'
                pv = ET.SubElement(pr0, 'ParameterValue', parameter=pid, format=fmt)
                ET.SubElement(pv, 'Data').text = str(val)
                ET.SubElement(pv, 'Qualifier', name='Variable Type').text = 'Dependent' 
            pa3 = ET.SubElement(pr0, 'ParameterValue', parameter='pa3', format='string')
            ET.SubElement(pa3, 'Data').text = 'Appearance'
            pa_set.update(('pa0', 'pa1', 'pa2', 'pa3'))
            pr_set.add('pr0')

        # pr1: Density
        if pd.notna(row.get("Density")):
            pr1 = ET.SubElement(bulk, 'PropertyData', property="pr1")
            ET.SubElement(pr1, 'Data', format='string').text = '-'
            ET.SubElement(pr1, 'Qualifier', name='Field Variable Compatible').text = 'Temperature'
            pa4 = ET.SubElement(pr1, 'ParameterValue', parameter='pa4', format='string')
            ET.SubElement(pa4, 'Data').text = 'Interpolation Options'
            ET.SubElement(pa4, 'Qualifier', name='AlgorithmType').text = 'Linear Multivariate'
            ET.SubElement(pa4, 'Qualifier', name='Normalized').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='Cached').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='ExtrapolationType').text = 'Projection to the Bounding Box'

            val, count = convert_unit(row.get("Density"), row.get("Density_units"))
            pa5 = ET.SubElement(pr1, 'ParameterValue', parameter='pa5', format='float')
            ET.SubElement(pa5, 'Data').text = str(val)
            ET.SubElement(pa5, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa_set.update(('pa4', 'pa5'))

            temp_val = row.get('Temperature')
            temp_unit = row.get('Temperature_units') or 'C'
            count = len(str(temp_val).split(','))
            pa6 = ET.SubElement(pr1, 'ParameterValue', parameter='pa6', format='float')
            ET.SubElement(pa6, 'Data').text = str(temp_val if pd.notna(temp_val) else '7.88860905221012e-31')
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            ET.SubElement(pa6, 'Qualifier', name='Field Variable').text = 'Temperature'
            ET.SubElement(pa6, 'Qualifier', name='Default Data').text = '22'
            ET.SubElement(pa6, 'Qualifier', name='Field Units').text = temp_unit
            ET.SubElement(pa6, 'Qualifier', name='Upper Limit').text = 'Program Controlled'
            ET.SubElement(pa6, 'Qualifier', name='Lower Limit').text = 'Program Controlled'
            pa_set.add('pa6')
            pr_set.add('pr1')

        # pr2: Elasticity
        if any(pd.notna(row.get(k)) for k in ["Youngs_Modulus_X", "Poissons_Ratio_XY", "Shear_Modulus_XY"]):
            pr2 = ET.SubElement(bulk, 'PropertyData', property="pr2")
            ET.SubElement(pr2, 'Data', format='string').text = '-'
            ET.SubElement(pr2, 'Qualifier', name='Behavior').text = 'Orthotropic'
            ET.SubElement(pr2, 'Qualifier', name='Field Variable Compatible').text = 'Temperature'
            # pa4
            pa4 = ET.SubElement(pr2, 'ParameterValue', parameter='pa4', format='string')
            ET.SubElement(pa4, 'Data').text = 'Interpolation Options'
            ET.SubElement(pa4, 'Qualifier', name='AlgorithmType').text = 'Linear Multivariate (Qhull)'
            ET.SubElement(pa4, 'Qualifier', name='Normalized').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='Cached').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='ExtrapolationType').text = 'Projection to the Bounding Box'
            pa_set.add('pa4')
            # pa23_31
            for pid, col, ucol in zip(
                ["pa23", "pa24", "pa25", "pa26", "pa27", "pa28", "pa29", "pa30", "pa31"], 
                ["Youngs_Modulus_X", "Youngs_Modulus_Y", "Youngs_Modulus_Z",
                    "Poissons_Ratio_XY", "Poissons_Ratio_YZ", "Poissons_Ratio_XZ",
                    "Shear_Modulus_XY", "Shear_Modulus_YZ", "Shear_Modulus_XZ"], 
                    ["Youngs_Modulus_units", "Youngs_Modulus_units", "Youngs_Modulus_units",
                 None, None, None, "Shear_Modulus_Unit", "Shear_Modulus_Unit", "Shear_Modulus_Unit"]):
                val, count = convert_unit(row.get(col), row.get(ucol) if ucol else None)
                param = ET.SubElement(pr2, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(param, 'Data').text = str(val)
                ET.SubElement(param, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa_set.update(('pa23', 'pa24', 'pa25', 'pa26', 'pa27', 'pa28', 'pa29', 'pa30', 'pa31'))
            # pa6
            pa6 = ET.SubElement(pr2, 'ParameterValue', parameter='pa6', format='float')
            temp_val = row.get('Temperature')
            count = len(str(temp_val).split(','))
            ET.SubElement(pa6, 'Data').text = str(temp_val if pd.notna(temp_val) else '7.88860905221012e-31')
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            ET.SubElement(pa6, 'Qualifier', name='Field Variable').text = 'Temperature'
            ET.SubElement(pa6, 'Qualifier', name='Default Data').text = '22'
            ET.SubElement(pa6, 'Qualifier', name='Field Units').text = row.get('Temperature_units') or 'C'
            ET.SubElement(pa6, 'Qualifier', name='Upper Limit').text = 'Program Controlled'
            ET.SubElement(pa6, 'Qualifier', name='Lower Limit').text = 'Program Controlled'
            pa_set.add('pa6')
            pr_set.add('pr2')


        if any(pd.notna(row.get(k)) for k in ["Youngs_Modulus", "Poissons_Ratio"]):
            pr2 = ET.SubElement(bulk, 'PropertyData', property="pr2")
            ET.SubElement(pr2, 'Data', format='string').text = '-'
            ET.SubElement(pr2, 'Qualifier', name='Behavior').text = 'Isotropic'
            ET.SubElement(pr2, 'Qualifier', name='Derive from').text = "Young's Modulus and Poisson's Ratio"
            ET.SubElement(pr2, 'Qualifier', name='Field Variable Compatible').text = 'Temperature'
            pa4 = ET.SubElement(pr2, 'ParameterValue', parameter='pa4', format='string')
            ET.SubElement(pa4, 'Data').text = 'Interpolation Options'
            ET.SubElement(pa4, 'Qualifier', name='AlgorithmType').text = 'Linear Multivariate'
            ET.SubElement(pa4, 'Qualifier', name='Normalized').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='Cached').text = 'True'
            ET.SubElement(pa4, 'Qualifier', name='ExtrapolationType').text = 'Projection to the Bounding Box'

            for pid, col, ucol in zip(["pa7", "pa8"], ["Youngs_Modulus", "Poissons_Ratio"], ["Youngs_Modulus_units", None]):
                val, count = convert_unit(row.get(col), row.get(ucol) if ucol else None)
                param = ET.SubElement(pr2, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(param, 'Data').text = str(val)
                ET.SubElement(param, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)

            pa6 = ET.SubElement(pr2, 'ParameterValue', parameter='pa6', format='float')
            temp_val = row.get('Temperature')
            count = len(str(temp_val).split(','))
            ET.SubElement(pa6, 'Data').text = str(temp_val if pd.notna(temp_val) else '7.88860905221012e-31')
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            ET.SubElement(pa6, 'Qualifier', name='Field Variable').text = 'Temperature'
            ET.SubElement(pa6, 'Qualifier', name='Default Data').text = '22'
            ET.SubElement(pa6, 'Qualifier', name='Field Units').text = row.get('Temperature_units') or 'C'
            ET.SubElement(pa6, 'Qualifier', name='Upper Limit').text = 'Program Controlled'
            ET.SubElement(pa6, 'Qualifier', name='Lower Limit').text = 'Program Controlled'
            pa_set.update(('pa6', 'pa7', 'pa8', 'pa4'))
            pr_set.add('pr2')

        # pr3: CTE
        if any(pd.notna(row.get(k)) for k in ['CTE_X', 'CTE_Y', 'CTE_Z']):
            pr3 = ET.SubElement(bulk, 'PropertyData', property="pr3")
            ET.SubElement(pr3, 'Data', format='string').text = '-'
            ET.SubElement(pr3, 'Qualifier', name='Definition').text = 'Instantaneous'
            ET.SubElement(pr3, 'Qualifier', name='Behavior').text = 'Orthotropic'
            # pa20_22
            for pid, col, ucol in zip(
                ["pa20", "pa21", "pa22"], 
                ["CTE_X", "CTE_Y", "CTE_Z"], 
                ["CTE_units", "CTE_units", "CTE_units"]):
                val, count = convert_unit(row.get(col), row.get(ucol) if ucol else None)
                param = ET.SubElement(pr3, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(param, 'Data').text = str(val)
                ET.SubElement(param, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa_set.update(('pa20', 'pa21', 'pa22'))
            # pa6
            pa6 = ET.SubElement(pr3, 'ParameterValue', parameter='pa6', format='float')
            temp_val = row.get('Temperature')
            count = len(str(temp_val).split(','))
            ET.SubElement(pa6, 'Data').text = str(temp_val if pd.notna(temp_val) else '7.88860905221012e-31')
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            pa_set.add('pa6')
            pr_set.add('pr3')

        if pd.notna(row.get("CTE")):
            pr3 = ET.SubElement(bulk, 'PropertyData', property="pr3")
            ET.SubElement(pr3, 'Data', format='string').text = '-'
            ET.SubElement(pr3, 'Qualifier', name='Definition').text = 'Instantaneous'
            ET.SubElement(pr3, 'Qualifier', name='Behavior').text = 'Isotropic'
            val, count = convert_unit(row.get("CTE"), row.get("CTE_units"))
            pa11 = ET.SubElement(pr3, 'ParameterValue', parameter='pa11', format='float')
            ET.SubElement(pa11, 'Data').text = str(val)
            ET.SubElement(pa11, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa6 = ET.SubElement(pr3, 'ParameterValue', parameter='pa6', format='float')
            temp_val = row.get('Temperature')
            count = len(str(temp_val).split(','))
            ET.SubElement(pa6, 'Data').text = str(temp_val if pd.notna(temp_val) else '7.88860905221012e-31')
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            pa_set.update(('pa6', 'pa11'))
            pr_set.add('pr3')

        # pr4: Damping
        if pd.notna(row.get('Damping_Ratio')):
            pr4 = ET.SubElement(bulk, 'PropertyData', property='pr4')
            ET.SubElement(pr4, 'Data', format='string').text = '-'
            for pid, col in zip(["pa12", "pa13"], ["Damping_Ratio", "CSDC"]):
                count = len(str(row.get(col)).split(','))
                pa12_13 = ET.SubElement(pr4, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(pa12_13, 'Data').text = str(row.get(col)) if pd.notna(row.get(col)) else '7.88860905221012e-31'
                ET.SubElement(pa12_13, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa_set.update(('pa12', 'pa13'))
            pr_set.add('pr4')

        # pr5: SN
        if any(pd.notna(row.get(k for k in ['A', 'm', 'C', 'r']))):  
            pr5 = ET.SubElement(bulk, 'PropertyData', property="pr5")
            ET.SubElement(pr5, 'Data', format='string').text = '-'
            ET.SubElement(pr5, 'Qualifier', name='Definition').text = 'Bilinear'
            ET.SubElement(pr5, 'Qualifier', name='Derive from').text = 'Coefficients and Exponents'
            for pid, col in zip(['pa14', 'pa15', 'pa16', 'pa17'], ['A', 'm', 'C', 'r']):
                count = len(str(row.get(col)).split(','))
                pa14_17 = ET.SubElement(pr5, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(pa14_17, 'Data').text = str(row.get(col))
                ET.SubElement(pa14_17, 'Qualifier', name='Variable Type').text = ",".join(Dep*count)
            pa_set.update(('pa14', 'pa15', 'pa16', 'pa17'))
            pr_set.add('pr5')

        # pr6: SS
        if any(pd.notna(row.get(k for k in ['Stress', 'Plastic_Strain']))):
            pr6 = ET.SubElement(bulk, 'PropertyData', property="pr6")
            ET.SubElement(pr6, 'Data', format='string').text = '-'
            ET.SubElement(pr6, 'Qualifier', name='Definition').text = 'Multilinear'
            for pid, col, ucol in zip(["pa18", "pa19"], ["Stress", "Plastic_Strain"], ["Stress_units", None]):
                val, count = convert_unit(row.get(col), row.get(ucol) if ucol else None)
                pa_18_19 = ET.SubElement(pr6, 'ParameterValue', parameter=pid, format='float')
                ET.SubElement(pa_18_19, 'Data').text = str(val)
                ET.SubElement(pa_18_19, 'Qualifier', name='Variable Type').text = ",".join(Dep*count) if col == "pa18" else ",".join(Indep*count)
            pa6 = ET.SubElement(pr6, 'ParameterValue', parameter='pa6', format='float')
            temp_val = row.get('Temperature')
            count = len(str(temp_val).split(','))
            ET.SubElement(pa6, 'Data').text = '7.88860905221012e-31' 
            ET.SubElement(pa6, 'Qualifier', name='Variable Type').text = ",".join(Indep*count)
            pa_set.update(('pa18', 'pa19', 'pa6'))
            pr_set.add('pr6')

    metadata = ET.SubElement(matml, 'Metadata')
    pa_info = {
        "pa0": ("Red", "Unitless"),
        "pa1": ("Green", "Unitless"),
        "pa2": ("Blue", "Unitless"),
        "pa3": ("Material Property", "Unitless"),
        "pa4": ("Options Variable", "Unitless"),
        "pa5": ("Density", "g/cm^3"),
        "pa6": ("Temperature", "C"),
        "pa7": ("Young's Modulus", "Pa"),
        "pa8": ("Poisson's Ratio", "Unitless"),
        "pa9": ("Bulk Modulus", "Pa"),
        "pa10": ("Shear Modulus", "Pa"),
        "pa11": ("Coefficient of Thermal Expansion", "1/C"),
        "pa12": ("Damping Ratio", "Unitless"),
        "pa13": ("Constant Structural Damping Coefficient", "Unitless"),
        "pa14": ("First Fatigue Strength Coefficient, A", "Unitless"),
        "pa15": ("First Fatigue Strength Exponent, m", "Unitless"),
        "pa16": ("Second Fatigue Strength Coefficient, C", "Unitless"),
        "pa17": ("Second Fatigue Strength Exponent, r", "Unitless"),
        "pa18": ("Stress", "Pa"),
        "pa19": ("Plastic Strain", "mm"),
        "pa20": ("Coefficient of Thermal Expansion X direction", "1/C"),
        "pa21": ("Coefficient of Thermal Expansion Y direction", "1/C"),
        "pa22": ("Coefficient of Thermal Expansion Z direction", "1/C"),
        "pa23": ("Young's Modulus X direction", "Pa"),
        "pa24": ("Young's Modulus Y direction", "Pa"),
        "pa25": ("Young's Modulus Z direction", "Pa"),
        "pa26": ("Poisson's Ratio XY", "Unitless"),
        "pa27": ("Poisson's Ratio YZ", "Unitless"),
        "pa28": ("Poisson's Ratio XZ", "Unitless"),
        "pa29": ("Shear Modulus XY", "Pa"),
        "pa30": ("Shear Modulus YZ", "Pa"),
        "pa31": ("Shear Modulus XZ", "Pa")
    }
    unit_groups = {
        "g/cm^3": ("Density", [("g", 1), ("cm", -3)]),
        "Pa": ("Stress", [("Pa", 1)]),
        "1/C": ("InvTemp1", [("C", -1)]),
        "C": ("Temperature", [("C", 1)]),
        "mm": ("Strain", [("mm", 1), ("mm", -1)]),
        "Unitless": (None, [])
    }

    for pa in sorted(pa_set):
        name, unit = pa_info.get(pa, (pa, "Unitless"))
        pd_elem = ET.SubElement(metadata, "ParameterDetails", id=pa)
        ET.SubElement(pd_elem, "Name").text = name
        if unit == "Unitless":
            ET.SubElement(pd_elem, "Unitless")
        else:
            group, units = unit_groups.get(unit, (None, []))
            if group:
                units_elem = ET.SubElement(pd_elem, "Units", name=group)
                for u_name, power in units:
                    u = ET.SubElement(units_elem, "Unit") if power == 1 else ET.SubElement(units_elem, "Unit", power=str(power))
                    ET.SubElement(u, "Name").text = u_name
            else:
                ET.SubElement(pd_elem, "Unitless")

    pr_info = {
        "pr0": "Color",
        "pr1": "Density",
        "pr2": "Elasticity",
        "pr3": "Coefficient of Thermal Expansion",
        "pr4": "Material Dependent Damping",
        "pr5": "S-N Curve",
        "pr6": "Isotropic Hardening"
    }
    for pr in sorted(pr_set):
        name = pr_info.get(pr, pr)
        pr_elem = ET.SubElement(metadata, "PropertyDetails", id=pr)
        ET.SubElement(pr_elem, "Unitless")
        ET.SubElement(pr_elem, "Name").text = name

    tree = ET.ElementTree(eng)
    try:
        ET.indent(tree, space="  ")
    except:
        pass
    tree.write(output_file, encoding='utf-8', xml_declaration=True)

if __name__ == '__main__':
    excel_to_ansys_fixed("xml.xlsx", "output_ansys.xml")